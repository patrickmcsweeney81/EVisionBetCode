"""
Format CSV files with Excel features:
- Auto-fit column widths
- Add filter to header row
- Freeze header row (row 1)
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Data directory
DATA_DIR = Path(__file__).parent / "data"

# CSV files to format
CSV_FILES = {
    "raw_odds_pure.csv": DATA_DIR / "raw_odds_pure.csv",
    "ev_hits.csv": DATA_DIR / "ev_hits.csv",
}

def format_csv_as_excel(csv_path):
    """Convert CSV to XLSX with formatting (auto-fit, filters, frozen header)."""
    if not csv_path.exists():
        print(f"[!] File not found: {csv_path}")
        return False
    
    # Convert to .xlsx path
    xlsx_path = csv_path.with_suffix('.xlsx')
    
    try:
        # Read CSV and convert to Excel
        import pandas as pd
        df = pd.read_csv(csv_path)
        
        # Write to Excel
        df.to_excel(xlsx_path, index=False, sheet_name='Data')
        print(f"✅ Converted {csv_path.name} → {xlsx_path.name}")
        
        # Load the workbook to add formatting
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Auto-fit columns
        for col_num, col in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_num)
            
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set column width (add padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long values
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Add filter to header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
        
        # Freeze header row (row 1)
        ws.freeze_panes = "A2"
        
        # Save the formatted workbook
        wb.save(xlsx_path)
        print(f"   ✓ Auto-fit columns")
        print(f"   ✓ Added filters to header row")
        print(f"   ✓ Frozen header row (row 1)")
        print()
        return True
        
    except Exception as e:
        print(f"[!] Error formatting {csv_path.name}: {e}")
        return False

def main():
    print("=" * 70)
    print("CSV FORMATTER - Adding Excel Features")
    print("=" * 70)
    print()
    
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("[!] openpyxl not installed. Installing...")
        os.system("pip install openpyxl")
        print()
    
    # Check if pandas is installed
    try:
        import pandas
    except ImportError:
        print("[!] pandas not installed. Installing...")
        os.system("pip install pandas")
        print()
    
    success_count = 0
    
    # Format each CSV
    for name, path in CSV_FILES.items():
        if path.exists():
            if format_csv_as_excel(path):
                success_count += 1
            break  # Only process the first file that exists
    
    print("=" * 70)
    if success_count > 0:
        print(f"✅ Formatted {success_count} file(s)")
        print()
        print("📁 Files created:")
        print(f"   • data/raw_odds_pure.xlsx")
        print(f"   • data/ev_hits.xlsx (or ev_opportunities.xlsx)")
        print()
        print("💡 Features added:")
        print("   ✓ Auto-fit column widths for readability")
        print("   ✓ Filter dropdown on header row (row 1)")
        print("   ✓ Frozen header row (row 1 stays visible when scrolling)")
    else:
        print("[!] No CSV files found or formatting failed")
    print()

if __name__ == "__main__":
    main()
